import openpyxl, requests

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1'
}

wb = openpyxl.load_workbook('21.xlsx') # change name for your file
sheet = wb['Sheet1']    # change according to your file
wb = openpyxl.Workbook()
new_sheet = wb['Sheet']
max = sheet.max_row
print(max)
url = ""
works = []
invalid_links = []
problematic_links = []
count_works = 0
count_invalid = 0
new_rows = 1
print("Идет проверка...")
for i in range(1, max + 1):
    url = sheet.cell(row=i, column=1).value
    try:
        response = requests.get(url, headers=headers)
        print(response.status_code)
    except requests.exceptions.RequestException as e:
        problematic_links.append(url)
        count_invalid += 1
        continue
    if response.status_code == 200:
        new_sheet.cell(row=new_rows, column=1).value = sheet.cell(row=i, column=1).value
        new_rows += 1
        works.append(url)
        count_works += 1
    else:
        invalid_links.append(url)
        count_invalid += 1
        continue
print(f"Проверено ссылок: {max}\nРабочие: {count_works}\nНерабочие: {count_invalid}")
print("Рабочие:")
for i in works:
    print(i)
print("Нерабочие:")
for i in invalid_links:
    print(i)
print("На ручную проверку:")
for i in problematic_links:
    print(i)
wb.save('working_links.xlsx')   # change to whatever name you would like to use


